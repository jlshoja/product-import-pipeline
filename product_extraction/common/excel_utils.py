"""
Shared Excel utility functions.

Provides reusable wrappers for common openpyxl / pandas Excel operations
and project-standard style constants used across tracker and report modules.

Unit 3 — Excel Operations Consolidation
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path


# ============================================================
# Excel I/O helpers
# ============================================================


def read_excel(filepath, **kwargs):
    """Read an Excel file into a DataFrame with openpyxl as default engine."""
    kwargs.setdefault("engine", "openpyxl")
    return pd.read_excel(filepath, **kwargs)


def write_dataframe(df, filepath, index=False, **kwargs):
    """Write a DataFrame to a single-sheet Excel file using openpyxl."""
    kwargs.setdefault("engine", "openpyxl")
    df.to_excel(filepath, index=index, **kwargs)


def excel_writer(filepath, **kwargs):
    """Return a pandas ExcelWriter context manager (openpyxl engine)."""
    kwargs.setdefault("engine", "openpyxl")
    return pd.ExcelWriter(filepath, **kwargs)


def load_workbook(filepath, **kwargs):
    """Load an openpyxl Workbook with error handling."""
    return openpyxl.load_workbook(filepath, **kwargs)


# ============================================================
# Project-standard style constants
# ============================================================

# --- Fills ---

GREEN_FILL = PatternFill("solid", start_color="C6EFCE")
RED_FILL = PatternFill("solid", start_color="FFC7CE")
YELLOW_FILL = PatternFill("solid", start_color="FFEB9C")
BLUE_FILL = PatternFill("solid", start_color="DDEBF7")
LIGHT_BLUE_FILL = PatternFill("solid", start_color="BDD7EE")

HEADER_FILL_DARK = PatternFill("solid", start_color="4472C4")   # blue header (compare_scans style)
HEADER_FILL_NONE = PatternFill(fill_type=None)                  # transparent header (report_generator style)
ROW_FILL_GREEN = PatternFill(fill_type='solid', fgColor='C6EFCE')

# --- Fonts ---

HEADER_FONT_DARK = Font(bold=True, color="FFFFFF", name="Arial")
HEADER_FONT_PLAIN = Font(bold=True, size=11)
BOLD_FONT = Font(bold=True, name="Arial")
NORMAL_FONT = Font(name="Arial")
RED_FONT = Font(bold=True, color="9C0006", name="Arial")
GREEN_FONT = Font(bold=True, color="276221", name="Arial")
CELL_FONT = Font(size=11)

# --- Borders ---

THIN_BORDER = Border(
    left=Side(border_style='thin'),
    right=Side(border_style='thin'),
    top=Side(border_style='thin'),
    bottom=Side(border_style='thin'),
)

# --- Alignments ---

CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER_TOP = Alignment(horizontal='center', vertical='top', readingOrder=2)
CENTER_CENTER = Alignment(horizontal='center', vertical='center', readingOrder=2)
RIGHT_CENTER = Alignment(horizontal='right', vertical='center', readingOrder=2)


# ============================================================
# Style helper functions
# ============================================================


def style_header(ws, row, cols, font=None, fill=None, alignment=None):
    """
    Apply standard header styling to a row.

    Args:
        ws:        openpyxl Worksheet
        row:       row number (1-based)
        cols:      number of columns to style
        font:      optional Font override (default: HEADER_FONT_DARK)
        fill:      optional PatternFill override (default: HEADER_FILL_DARK)
        alignment: optional Alignment override (default: CENTER_WRAP)
    """
    if font is None:
        font = HEADER_FONT_DARK
    if fill is None:
        fill = HEADER_FILL_DARK
    if alignment is None:
        alignment = CENTER_WRAP

    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font
        cell.fill = fill
        cell.alignment = alignment


def auto_width(ws, max_width=50):
    """
    Auto-size column widths based on content length.

    Args:
        ws:        openpyxl Worksheet
        max_width: maximum column width cap (default: 50)
    """
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)


def set_fill(ws, row, cols, fill):
    """
    Apply a PatternFill to an entire row range.

    Args:
        ws:   openpyxl Worksheet
        row:  row number (1-based)
        cols: number of columns to fill
        fill: PatternFill instance
    """
    for col in range(1, cols + 1):
        ws.cell(row=row, column=col).fill = fill
