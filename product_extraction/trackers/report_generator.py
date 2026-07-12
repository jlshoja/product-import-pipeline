#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Report Generator - Generate HTML and Excel reports
"""

import sys
from pathlib import Path
from typing import List, Dict, Optional
from datetime import datetime

import pandas as pd
from openpyxl import Workbook

# Enable direct-script execution
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

# ─── Shared Excel utilities (Unit 3) ───────────────────────────────
from common.excel_utils import (
    HEADER_FILL_NONE,
    ROW_FILL_GREEN,
    THIN_BORDER,
    HEADER_FONT_PLAIN,
    CELL_FONT,
    CENTER_TOP,
    CENTER_CENTER,
    RIGHT_CENTER,
    excel_writer,
)
from common.file_utils import ensure_directory

COL_WIDTHS = {'A': 3.44, 'B': 18.22, 'C': 37.11}


def _apply_extracted_style(ws, rows: List[Dict], columns: List[str]) -> None:
    """
    Write data into worksheet with the same style as extracted_products.xlsx
    
    columns: list of keys from each row dict that map to B, C, ... columns
             (column A is always the row number)
    """
    # ---- Header row ----
    headers = ['No'] + columns
    col_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G']

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT_PLAIN
        cell.alignment = CENTER_TOP
        cell.border    = THIN_BORDER

    # ---- Data rows ----
    for row_num, item in enumerate(rows, start=1):
        # Column A — row number
        cell_a = ws.cell(row=row_num + 1, column=1, value=row_num)
        cell_a.font      = CELL_FONT
        cell_a.alignment = CENTER_CENTER

        # Remaining columns
        for col_idx, key in enumerate(columns, start=2):
            cell = ws.cell(row=row_num + 1, column=col_idx, value=item.get(key, ''))
            cell.font      = CELL_FONT
            cell.fill      = ROW_FILL_GREEN
            cell.alignment = CENTER_CENTER

    # ---- Column widths ----
    dynamic_widths = {'A': 3.44, 'B': 18.22, 'C': 37.11}
    for letter, width in dynamic_widths.items():
        ws.column_dimensions[letter].width = width


def generate_new_products_excel(
    new_products: List[Dict],
    output_dir: Path,
    date_str: Optional[str] = None
) -> Optional[Path]:
    """
    Generate extracted-style Excel file for new products.

    Args:
        new_products : list of dicts, each must contain at minimum:
                       'product_name' (or 'name') and 'product_url' (or 'url')
        output_dir   : directory where the file will be saved
        date_str     : optional date suffix (default: today YYYY-MM-DD)

    Returns:
        Path to created file, or None on failure.

    Output filename example:
        new_products_2024-03-15.xlsx
    """
    if not new_products:
        print("[INFO] No new products — skipping new_products file.")
        return None

    try:
        output_dir = Path(output_dir)
        ensure_directory(output_dir)

        date_str = date_str or datetime.now().strftime('%Y-%m-%d')
        output_path = output_dir / f'extracted_products_new_{date_str}.xlsx'

        # Normalise field names
        normalised = []
        for p in new_products:
            normalised.append({
                'Product Name': p.get('product_name') or p.get('name', ''),
                'Product URL':  p.get('product_url')  or p.get('url')          or p.get('link', ''),
            })

        wb = Workbook()
        ws = wb.active
        ws.title = 'New Products'
        ws.sheet_view.rightToLeft = True

        _apply_extracted_style(ws, normalised, ['Product Name', 'Product URL'])

        wb.save(output_path)
        print(f"[OK] New products file saved: {output_path.name}  ({len(normalised)} products)")
        return output_path

    except Exception as e:
        print(f"[ERROR] Could not generate new_products file: {e}")
        return None


def generate_price_changes_excel(
    price_changes: List[Dict],
    output_dir: Path,
    date_str: Optional[str] = None
) -> Optional[Path]:
    """
    Generate extracted-style Excel file for products with price changes.

    Args:
        price_changes : list of dicts, each must contain at minimum:
                        'product_name' (or 'name') and 'product_url' (or 'url')
        output_dir    : directory where the file will be saved
        date_str      : optional date suffix (default: today YYYY-MM-DD)

    Returns:
        Path to created file, or None on failure.

    Output filename example:
        price_changes_2024-03-15.xlsx
    """
    if not price_changes:
        print("[INFO] No price changes — skipping price_changes file.")
        return None

    try:
        output_dir = Path(output_dir)
        ensure_directory(output_dir)

        date_str = date_str or datetime.now().strftime('%Y-%m-%d')
        output_path = output_dir / f'extracted_products_changes_{date_str}.xlsx'

        # Normalise field names
        normalised = []
        for p in price_changes:
            normalised.append({
                'Product Name': p.get('product_name') or p.get('name', ''),
                'Product URL':  p.get('product_url')  or p.get('url')          or p.get('link', ''),
            })

        wb = Workbook()
        ws = wb.active
        ws.title = 'Price Changes'
        ws.sheet_view.rightToLeft = True

        _apply_extracted_style(ws, normalised, ['Product Name', 'Product URL'])

        wb.save(output_path)
        print(f"[OK] Price changes file saved: {output_path.name}  ({len(normalised)} products)")
        return output_path

    except Exception as e:
        print(f"[ERROR] Could not generate price_changes file: {e}")
        return None


# ============================================================
# Original functions (unchanged)
# ============================================================

def generate_html_report(
    current_df: pd.DataFrame,
    new_products: List[Dict],
    price_changes: List[Dict],
    removed_products: List[Dict],
    output_path: Path
) -> bool:
    """
    Generate HTML report

    Args:
        current_df: Current products DataFrame
        new_products: List of new products
        price_changes: List of price changes
        removed_products: List of removed products
        output_path: Output file path

    Returns:
        True if successful
    """
    try:
        html_content = f"""
<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
    <meta charset="UTF-8">
    <title>Product Report</title>
    <style>
        body {{ font-family: Tahoma; padding: 20px; }}
        h1 {{ color: #333; }}
        table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: right; }}
        th {{ background-color: #4CAF50; color: white; }}
        tr:nth-child(even) {{ background-color: #f2f2f2; }}
    </style>
</head>
<body>
    <h1>Product Tracking Report</h1>
    <p>Total Products: {len(current_df)}</p>
    <p>New Products: {len(new_products)}</p>
    <p>Price Changes: {len(price_changes)}</p>
    <p>Removed Products: {len(removed_products)}</p>

    <h2>Current Products</h2>
    {current_df.to_html(index=False, classes='dataframe')}
</body>
</html>
"""
        ensure_directory(output_path.parent)

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

        return True

    except Exception as e:
        print(f"Error generating HTML report: {e}")
        return False


def generate_excel_report(
    current_df: pd.DataFrame,
    new_products: List[Dict],
    price_changes: List[Dict],
    removed_products: List[Dict],
    output_path: Path
) -> bool:
    """
    Generate Excel report with multiple sheets

    Args:
        current_df: Current products DataFrame
        new_products: List of new products
        price_changes: List of price changes
        removed_products: List of removed products
        output_path: Output file path

    Returns:
        True if successful
    """
    try:
        ensure_directory(output_path.parent)

        with excel_writer(output_path) as writer:
            current_df.to_excel(writer, sheet_name='Current Products', index=False)

            if new_products:
                pd.DataFrame(new_products).to_excel(
                    writer, sheet_name='New Products', index=False
                )

            if price_changes:
                pd.DataFrame(price_changes).to_excel(
                    writer, sheet_name='Price Changes', index=False
                )

            if removed_products:
                pd.DataFrame(removed_products).to_excel(
                    writer, sheet_name='Removed Products', index=False
                )

        return True

    except Exception as e:
        print(f"Error generating Excel report: {e}")
        return False


if __name__ == "__main__":
    # Quick smoke-test with dummy data
    sample_new = [
        {'product_name': 'کیف زنانه کد 1001', 'product_url': 'https://example.com/1001'},
        {'product_name': 'کیف مردانه کد 2002', 'product_url': 'https://example.com/2002'},
    ]
    sample_changes = [
        {'product_name': 'کیف پاسپورتی کد 3003', 'product_url': 'https://example.com/3003'},
    ]

    out = Path('/tmp/test_reports')
    generate_new_products_excel(sample_new, out)
    generate_price_changes_excel(sample_changes, out)
    print("Done — check /tmp/test_reports/")
